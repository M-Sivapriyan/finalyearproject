from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import sqlite3

conn = sqlite3.connect('FaceBase.db')
c = conn.cursor()
wb = Workbook()
dest_filename = 'DataBase.xlsx'
cmd="SELECT * FROM Person ORDER BY RollNo"
c.execute(cmd)
ws1 = wb.active
ws1.title = "Attendance"
ws1.append(('RollNo','Name'))
ws1.append(('',''))
while True:
    a = c.fetchone()
    if a == None:
        break
    else:
        ws1.append((a[1],a[2]))

wb.save(filename = "DataBase.xlsx")
wb.save(filename = dest_filename)
    
    
