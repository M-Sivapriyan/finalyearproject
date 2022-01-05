from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import sqlite3

conn = sqlite3.connect('FaceBase.db')
c = conn.cursor()

wb = Workbook()
dest_filename = 'Attendance.xlsx'
cmd="SELECT * FROM Attendance ORDER BY RollNo"
c.execute(cmd)

ws1 = wb.active
ws1.title = "Attendance"
ws1.append(('RollNo','Name','Date','Time'))
ws1.append(('', '', '',''))

while True:
    a = c.fetchone()
    if a == None:
        break
    else:
        ws1.append((a[0], a[1],a[2],a[3]))

    #saving the file
wb.save(filename = "Attendance.xlsx")
wb.save(filename = dest_filename)
    
    
